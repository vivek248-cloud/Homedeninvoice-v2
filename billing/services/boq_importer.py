from openpyxl import load_workbook


class BOQImporter:

    def __init__(self, file):
        self.file = file

    def parse(self):

        wb = load_workbook(self.file)
        ws = wb.active

        rows = []

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            rows.append({
                "floor": row[0] or "",
                "room": row[1] or "",
                "element": row[2] or "",
                "series": row[3] or "",
                "length": row[4] or 0,
                "width": row[5] or 0,
                "qty": row[6] or 1,
                "rate": row[7] or 0,
            })

        return rows